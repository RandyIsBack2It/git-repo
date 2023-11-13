import openpyxl
import re

# Function for performing a keyword search in an Excel file
def keyword_search():
    file_name = "study_tool.xlsx"  # Define the Excel file name

    keyword = input("Enter the keyword to search for: ")  # Get user input for the search keyword
    print()  # Add a blank line here
    pattern = r'\b' + re.escape(keyword) + r'\b'  # Prepare a regular expression pattern for the whole word match

    try:
        workbook = openpyxl.load_workbook(file_name)  # Open the Excel file
        sheet = workbook.active
        matching_rows = []  # Initialize a list to store rows with matching data
        headers = None  # Initialize variable to store column headers

        for row in sheet.iter_rows(values_only=True):
            if not headers:
                headers = row  # Store the first row as headers
                continue

            if any(cell is not None for cell in row):
                for cell in row:
                    if cell is not None and re.search(pattern, str(cell), re.I):
                        matching_rows.append(row)  # Add the row to matching rows if it contains the keyword
                        break

        if matching_rows:
            # Display matching rows with column headers
            for row in matching_rows:
                if any(cell is not None for cell in row):
                    for header, cell in zip(headers, row):
                        if cell is not None:
                            print(f"{header}: {cell}")
                    print()
            print()  # Add a blank line here
        else:
            print("No matching rows found for the keyword.")
        
        run_again = input("Do you want to run Keyword Search again? (y/N): ").strip().lower()
        if run_again == 'y':
            print()  # Add a line break before rerunning the function
            keyword_search()
        else:
            return  # Return to the main menu

        workbook.close()  # Close the Excel file

    except FileNotFoundError:
        print(f"File '{file_name}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Function for spellcheck feature (Option 2)
def spellcheck():
    file_name = "study_tool.xlsx"  # Define the Excel file name

    keyword = input("Enter a keyword to find matching words: ").lower()
    print()  # Add a blank line here

    try:
        workbook = openpyxl.load_workbook(file_name)  # Open the Excel file
        sheet = workbook.active

        words = set()  # Use a set to store unique words
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    words_in_cell = cell.split()  # Split the cell content into individual words
                    for word in words_in_cell:
                        word = word.strip(",.!?")  # Remove common punctuation marks
                        if len(word) >= 2 and word[0].lower() == keyword[0] and word[-1].lower() == keyword[-1]:
                            words.add(word.lower())  # Store words in lowercase to ensure uniqueness

        if words:
            # Sort the words by the number of common letters
            words = sorted(words, key=lambda x: sum(1 for a, b in zip(x, keyword) if a == b), reverse=True)

            # Display the words
            print("\nMatching words (sorted by common letters):")
            for word in words:
                print(word)
            print()
        else:
            print("No matching words found for the given criteria.")
        
        run_again = input("Do you want to run Spellcheck again? (y/N): ").strip().lower()
        if run_again == 'y':
            print()  # Add a line break before rerunning the function
            spellcheck()
        else:
            return  # Return to the main menu

        workbook.close()  # Close the Excel file

    except FileNotFoundError:
        print(f"File '{file_name}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Main menu loop
while True:
    print("\nSelect an Option:")
    print("  1. Keyword Search")
    print("  2. Spellcheck")
    print("  3. Exit (Default)")
    print()  # Add a blank line
    choice = input("Select an option (1/2/3): ").strip() or '3'  # Get the user's choice, default to '3' if Enter is pressed
    print()  # Add a blank line here

    if choice == '1':
        keyword_search()  # Call the keyword_search function for option 1
    elif choice == '2':
        spellcheck()  # Call the spellcheck function for option 2
    elif choice == '3':
        print("Exiting the script. Goodbye!")  # Display a goodbye message
        print()  # Add a blank line here
        break
    else:
        print
